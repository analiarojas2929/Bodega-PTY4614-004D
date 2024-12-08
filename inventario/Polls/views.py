from datetime import datetime
import os
import requests
import json
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Material
from .serializers import MaterialSerializer
import openpyxl
from .utils import format_date
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from datetime import datetime
from django.http import HttpResponse, JsonResponse
from django.db import models
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.utils.timezone import make_aware
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.db.models import Sum, F
from .models import Material, Ticket, CustomUser, Role, Question, Choice
from .forms import CustomUserForm, TicketForm, MaterialForm
from .serializers import QuestionSerializer, ChoiceSerializer
from .roles import ADMINISTRADOR_SISTEMA, ADMINISTRADOR_OBRA, JEFE_OBRA, CAPATAZ, JEFE_BODEGA

from rest_framework import viewsets
from django.contrib.auth.forms import AuthenticationForm

API_BASE_URL = "http://127.0.0.1:8000/api"
JSON_FILE_PATH = os.path.join(settings.BASE_DIR, 'api', 'materiales_data.json')


def has_role_id(user, role_id):
    return user.roles.filter(id=role_id).exists()

# Vista principal
def home_view(request):
    context = get_role_context(request.user)
    total_materiales = Material.objects.count()
    materiales_disponibles = Material.objects.aggregate(total=Sum('cantidad_disponible'))['total'] or 0
    materiales_en_movimiento = Ticket.objects.filter(estado='pendiente').aggregate(total=Sum('cantidad'))['total'] or 0

    # Obtener los movimientos recientes (últimos 5)
    movimientos_recientes = Ticket.objects.order_by('-fecha_creacion')[:5]

    # Obtener alertas de stock bajo
    alertas_stock = Material.objects.filter(cantidad_disponible__lt=F('stock')).order_by('cantidad_disponible')[:5]

    # Obtener solicitudes recientes
    solicitudes_recientes = Ticket.objects.filter(estado='pendiente').order_by('-fecha_creacion')[:5]

    context = {
        'total_materiales': total_materiales,
        'materiales_disponibles': materiales_disponibles,
        'materiales_en_movimiento': materiales_en_movimiento,
        'movimientos_recientes': movimientos_recientes,
        'alertas_stock': alertas_stock,
        'solicitudes_recientes': solicitudes_recientes,
    }
    return render(request, 'Modulo_usuario/HomeView/home.html', context)

def get_role_context(user):
    return {
        'is_jefe_bodega': user.roles.filter(id=JEFE_BODEGA).exists(),
        'is_jefe_obra': user.roles.filter(id=JEFE_OBRA).exists(),
        'is_capataz': user.roles.filter(id=CAPATAZ).exists(),
        'can_access_ticket': user.roles.filter(id=CAPATAZ).exists() or user.roles.filter(id=JEFE_OBRA).exists(),
        'can_access_inventario': user.roles.filter(id=JEFE_BODEGA).exists() or user.roles.filter(id=JEFE_OBRA).exists(),
        'can_access_reportes': user.roles.filter(id=ADMINISTRADOR_OBRA).exists(),
        'is_administrador_obra': user.roles.filter(id=ADMINISTRADOR_OBRA).exists(),
        'is_administrador_sistema': user.roles.filter(id=ADMINISTRADOR_SISTEMA).exists(),
        'can_access_list_ticket': user.roles.filter(id=JEFE_BODEGA).exists(),
    }


# Vista de inventario (solo accesible por Jefe de Bodega, Role ID = 5)
@login_required
@user_passes_test(lambda u: has_role_id(u, JEFE_BODEGA) or has_role_id(u, JEFE_OBRA), login_url='/access_denied/')
def inventory(request):
    context = get_role_context(request.user)
    return render(request, 'Modulo_usuario/InventoryView/inventory.html', context)


# Obtener la ruta base del proyecto
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
@login_required
@user_passes_test(lambda u: has_role_id(u, JEFE_BODEGA) or has_role_id(u, JEFE_OBRA), login_url='/access_denied/')
def lista_view(request):
    """
    Vista para listar materiales obtenidos desde la API
    """
    api_url = f"{settings.API_BASE_URL}/materiales/"
    materiales_json = []

    try:
        # Realizar una petición GET a la API para obtener la lista de materiales
        response = requests.get(api_url)
        if response.status_code == 200:
            materiales_json = response.json()
        else:
            messages.error(request, f"Error al obtener datos de la API: {response.status_code} - {response.text}")

    except requests.exceptions.RequestException as e:
        messages.error(request, f"Error al conectar con la API: {e}")

    return render(request, 'Modulo_usuario/InventoryView/lista.html', {
        'materiales_json': materiales_json
    })
@login_required(login_url='/admin_login/')
@user_passes_test(lambda u: has_role_id(u, ADMINISTRADOR_SISTEMA), login_url='/access_denied/')
def home_admin(request):
    return render(request, 'Modulo_administrador/usuarios/home_admin.html')


@login_required(login_url='/admin_login/')
def restricted_view(request):
    usuario = request.user
    roles = usuario.roles.all()  # Obtener los roles del usuario

    context = {
        'usuario': usuario,
        'roles': roles,
        'is_administrador_obra': has_role_id(usuario, ADMINISTRADOR_OBRA),
        'is_jefe_bodega': has_role_id(usuario, JEFE_BODEGA),
        'is_jefe_obra': has_role_id(usuario, JEFE_OBRA),
        'is_administrador_sistema': has_role_id(usuario, ADMINISTRADOR_SISTEMA),
    }
    return render(request, 'Modulo_administrador/usuarios/restricted_view.html', context)



@login_required(login_url='/admin_login/')
@user_passes_test(lambda u: has_role_id(u, ADMINISTRADOR_SISTEMA), login_url='/access_denied/')
def redirect_home_administrador(request):
    """
    Vista restringida que muestra detalles del usuario autenticado.
    Solo usuarios autenticados pueden acceder a esta vista.
    """
    usuario = request.user
    roles = usuario.roles.all()  # Obtener los roles del usuario

    # Depuración: Imprimir roles en la consola (opcional)
    print(f"Usuario: {usuario.username}, Roles: {roles}")

    context = {
        'usuario': usuario,
        'roles': roles,
        'mensaje': 'Bienvenido a la vista restringida'
    }
    return render(request, 'Modulo_administrador/usuarios/restricted_view.html', context)

# Restaurar material inactivo (solo accesible por Jefe de Bodega)
@login_required
@user_passes_test(lambda u: has_role_id(u, JEFE_BODEGA), login_url='/access_denied/')
def restore_material_view(request, id):
    material = get_object_or_404(Material, id=id)
    if request.method == 'POST':
        material.activo = True
        material.save()
        return redirect('lista_view')
    return render(request, 'Modulo_usuario/InventoryView/restaurar.html', {'material': material})


# Agregar material (solo accesible por Jefe de Bodega)


def add_material_view(request):
    """
    Vista para agregar un material y guardar en la base de datos y JSON.
    """
    if not os.path.exists(JSON_FILE_PATH):
        # Crear el archivo JSON si no existe
        with open(JSON_FILE_PATH, 'w', encoding='utf-8') as file:
            json.dump([], file, ensure_ascii=False, indent=4)

    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Guardar el material en la base de datos
                    nuevo_material = form.save()

                    # Crear la estructura de datos para JSON
                    material_data = {
                        "id": nuevo_material.id,
                        "nombre": nuevo_material.nombre,
                        "descripcion": nuevo_material.descripcion,
                        "unidad_medida": nuevo_material.unidad_medida.descripcion,
                        "cantidad_disponible": nuevo_material.cantidad_disponible,
                        "stock": nuevo_material.stock,
                        "activo": nuevo_material.activo
                    }

                    # Leer el archivo JSON existente
                    with open(JSON_FILE_PATH, 'r', encoding='utf-8') as file:
                        materiales_json = json.load(file)

                    # Agregar el nuevo material
                    materiales_json.append(material_data)

                    # Escribir los datos actualizados al archivo JSON
                    with open(JSON_FILE_PATH, 'w', encoding='utf-8') as file:
                        json.dump(materiales_json, file, ensure_ascii=False, indent=4)

                messages.success(request, "Material agregado correctamente y guardado en JSON.")
                return redirect('lista_view')

            except Exception as e:
                form.add_error(None, f"Error al guardar el material: {str(e)}")
    else:
        form = MaterialForm()

    return render(request, 'Modulo_usuario/InventoryView/agregar.html', {'form': form})
    
def actualizar_stock(material_id, nueva_cantidad):
    url = f"{settings.API_BASE_URL}/materiales/{material_id}/"
    data = {'cantidad_disponible': nueva_cantidad}
    response = requests.patch(url, json=data)
    return response.status_code == 200

# Actualizar material (solo accesible por Jefe de Bodega)
# Ruta al archivo JSON
@login_required
@user_passes_test(lambda u: has_role_id(u, JEFE_BODEGA), login_url='/access_denied/')
def editar_material(request, material_id):
    # Obtener el material por ID o retornar un error 404 si no existe
    material = get_object_or_404(Material, id=material_id)

    # Inicializar el formulario con los datos existentes del material
    if request.method == 'POST':
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            try:
                # Actualizar en la base de datos local
                updated_material = form.save()

                # Preparar los datos para la API
                api_url = f"{settings.API_BASE_URL}/materiales/{material_id}/"
                payload = {
                    "nombre": updated_material.nombre,
                    "descripcion": updated_material.descripcion,
                    "unidad_medida": updated_material.unidad_medida.id,  # Enviar ID de unidad de medida
                    "cantidad_disponible": updated_material.cantidad_disponible,
                    "stock": updated_material.stock,
                    "activo": updated_material.activo
                }

                headers = {'Content-Type': 'application/json'}

                # Realizar la solicitud PATCH a la API
                response = requests.patch(api_url, json=payload, headers=headers)

                if response.status_code == 200:
                    # Actualizar el archivo JSON local
                    actualizar_json_material(updated_material)
                    messages.success(request, 'Material actualizado correctamente en todos los sistemas.')
                else:
                    # Si la API devuelve un error, mostrar el mensaje
                    messages.warning(request, f"Material actualizado localmente, pero hubo un error al actualizar en la API: {response.status_code} - {response.text}")

                return redirect('lista_view')

            except Exception as e:
                # Manejar cualquier excepción que ocurra durante el proceso
                messages.error(request, f"Error al actualizar el material: {str(e)}")
        else:
            messages.error(request, f"Errores en el formulario: {form.errors}")

    else:
        form = MaterialForm(instance=material)

    return render(request, 'Modulo_usuario/InventoryView/editar.html', {
        'form': form,
        'material': material
    })



# Eliminar material (solo accesible por Jefe de Bodega)
@login_required
@user_passes_test(lambda u: has_role_id(u, JEFE_BODEGA), login_url='/access_denied/')
def delete_material_view(request, id):
    """
    Vista para eliminar definitivamente un material tanto de la base de datos como del archivo JSON.
    """
    # Obtener el material de la base de datos
    material = get_object_or_404(Material, id=id)
    
    if request.method == 'POST':
        # Eliminar el material de la base de datos
        material.delete()

        # Eliminar del archivo JSON
        try:
            # Leer el archivo JSON existente
            if os.path.exists(JSON_FILE_PATH):
                with open(JSON_FILE_PATH, 'r', encoding='utf-8') as file:
                    materiales_json = json.load(file)
                
                # Filtrar el material que queremos eliminar
                materiales_json = [m for m in materiales_json if m['id'] != id]

                # Guardar la lista actualizada en el archivo JSON
                with open(JSON_FILE_PATH, 'w', encoding='utf-8') as file:
                    json.dump(materiales_json, file, ensure_ascii=False, indent=4)
            
            messages.success(request, "Material eliminado definitivamente.")
        except (FileNotFoundError, json.JSONDecodeError) as e:
            messages.error(request, f"Error al actualizar el archivo JSON: {e}")

        return redirect('lista_view')

    return render(request, 'Modulo_usuario/InventoryView/eliminar.html', {'material': material})
# Crear ticket (solo accesible por Jefe de Obra  Jefe de Bodega)
@login_required
@user_passes_test(lambda u: has_role_id(u, JEFE_OBRA) or has_role_id(u, CAPATAZ), login_url='/access_denied/')
def crear_ticket(request):
    try:
        # Obtener la lista de materiales desde la API
        response = requests.get(f'{settings.API_BASE_URL}/materiales/')
        materiales_json = response.json() if response.status_code == 200 else []
    except requests.exceptions.RequestException:
        materiales_json = []
        messages.error(request, "Error al obtener la lista de materiales desde la API.")

    if request.method == 'POST':
        usuario = request.user
        try:
            # Obtener la lista de materiales del formulario
            materiales_data = json.loads(request.body).get('materiales', [])

            # Validar que haya materiales seleccionados
            if not materiales_data:
                messages.error(request, 'No se han seleccionado materiales.')
                return JsonResponse({'success': False}, status=400)

            for material_data in materiales_data:
                nombre_material = material_data.get('nombre')
                cantidad = int(material_data.get('cantidad'))
                estado = material_data.get('estado')

                # Validar que el material existe en la base de datos
                material = Material.objects.filter(nombre=nombre_material).first()
                if not material:
                    messages.error(request, f'El material "{nombre_material}" no existe.')
                    continue

                # Validar si hay suficiente stock disponible
                if material.cantidad_disponible < cantidad and estado == 'cobrado':
                    messages.error(request, f'No hay suficiente stock para "{nombre_material}".')
                    continue

                # Crear el ticket
                Ticket.objects.create(
                    usuario=usuario,
                    material_solicitado=material,
                    cantidad=cantidad,
                    estado=estado
                )

                # Descontar stock y actualizar JSON si es necesario
                if estado == 'cobrado':
                    material.cantidad_disponible -= cantidad
                    material.save()
                    actualizar_json_material(material)  # Actualizar el archivo JSON

            messages.success(request, 'Ticket creado exitosamente.')
            return JsonResponse({'success': True})

        except Exception as e:
            messages.error(request, f'Error al crear el ticket: {str(e)}')
            return JsonResponse({'success': False}, status=500)

    return render(request, 'Modulo_usuario/tickets/crear.html', {'materiales': materiales_json})


# Lista de tickets (solo accesible por Jefe de Obra o Jefe de Bodega)
def has_any_role(user, roles):
    return any(user.roles.filter(id=role).exists() for role in roles)

@login_required
@user_passes_test(lambda u: has_any_role(u, [JEFE_OBRA, JEFE_BODEGA, CAPATAZ]), login_url='/access_denied/')
def lista_tickets(request):
    tickets = Ticket.objects.all()

    # Filtrar por estado
    estado = request.GET.get('estado')
    if estado:
        tickets = tickets.filter(estado=estado)
    
    return render(request, 'Modulo_usuario/tickets/lista.html', {'tickets': tickets})

def es_ajax(request):
    """Verifica si la solicitud es AJAX."""
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'

@login_required
@user_passes_test(lambda u: has_role_id(u, JEFE_OBRA) or has_role_id(u, JEFE_BODEGA), login_url='/access_denied/')
def cobrar_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    material = ticket.material_solicitado

    # Verificar si el ticket ya fue cobrado
    if ticket.estado == 'cobrado':
        messages.warning(request, "El ticket ya ha sido cobrado anteriormente.")
        return redirect('lista_tickets')

    # Verificar si hay suficiente stock disponible
    if material.cantidad_disponible < ticket.cantidad:
        messages.error(request, 'No hay suficiente stock disponible para cobrar este ticket.')
        return redirect('lista_tickets')

    try:
        with transaction.atomic():
            # Descontar stock
            print(f"Stock actual de {material.nombre}: {material.cantidad_disponible}")
            material.cantidad_disponible -= ticket.cantidad
            material.save()
            print(f"Nuevo stock de {material.nombre}: {material.cantidad_disponible}")

            # Cambiar estado del ticket
            ticket.estado = 'cobrado'
            ticket.save(update_fields=['estado'])

            # Actualizar archivo JSON
            if actualizar_json_material(material):
                print(f"Material {material.nombre} actualizado en JSON correctamente.")
            else:
                print(f"[ERROR] No se pudo actualizar el material {material.nombre} en el JSON.")

        messages.success(request, "Ticket cobrado y stock actualizado correctamente.")

    except Exception as e:
        print(f"[ERROR] Error al cobrar el ticket: {e}")
        messages.error(request, f"Error al cobrar el ticket: {e}")

    return redirect('lista_tickets')

def actualizar_json_material(material):
    try:
        if not os.path.exists(JSON_FILE_PATH):
            print(f"[ERROR] Archivo JSON no encontrado: {JSON_FILE_PATH}")
            return False

        with open(JSON_FILE_PATH, 'r+', encoding='utf-8') as file:
            try:
                materiales_json = json.load(file)
            except json.JSONDecodeError:
                print(f"[ERROR] Archivo JSON mal formado.")
                return False

            # Buscar y actualizar el material en el JSON
            material_actualizado = False
            for m in materiales_json:
                if m['id'] == material.id:
                    m.update({
                        "nombre": material.nombre,
                        "descripcion": material.descripcion,
                        "unidad_medida": material.unidad_medida.descripcion,
                        "cantidad_disponible": material.cantidad_disponible,
                        "stock": material.stock,
                        "activo": material.activo,
                    })
                    material_actualizado = True
                    break

            if not material_actualizado:
                print(f"[INFO] Material {material.nombre} no encontrado en JSON. Agregando nuevo.")
                materiales_json.append({
                    "id": material.id,
                    "nombre": material.nombre,
                    "descripcion": material.descripcion,
                    "unidad_medida": material.unidad_medida.descripcion,
                    "cantidad_disponible": material.cantidad_disponible,
                    "stock": material.stock,
                    "activo": material.activo,
                })

            # Sobrescribir el archivo JSON
            file.seek(0)
            json.dump(materiales_json, file, ensure_ascii=False, indent=4)
            file.truncate()

        return True

    except Exception as e:
        print(f"[ERROR] Error al actualizar el archivo JSON: {e}")
        return False

# Ver ticket (solo accesible por Jefe de Obra o Jefe de Bodega)
@login_required
@user_passes_test(lambda u: has_role_id(u, JEFE_OBRA) or has_role_id(u, JEFE_BODEGA), login_url='/access_denied/')
def ver_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    return render(request, 'Modulo_usuario/tickets/ver.html', {'ticket': ticket})


# Eliminar ticket (solo accesible por Jefe de Obra o Jefe de Bodega)
@login_required
def eliminar_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    try:
        with transaction.atomic():
            # Remove the line that reverts the material quantity
            ticket.delete()
            messages.success(request, "Ticket eliminado correctamente.")
    except Exception as e:
        messages.error(request, f"Error al eliminar el ticket: {e}")

    return redirect('lista_tickets')

# Vista para alertas de stock (solo accesible por Administrador de Obra)
@login_required
@user_passes_test(lambda u: has_role_id(u, ADMINISTRADOR_OBRA), login_url='/access_denied/')
def stock_alerts_view(request):
    alertas_stock = Material.objects.filter(cantidad_disponible__lt=models.F('stock'))
    return render(request, 'Modulo_usuario/ReportsView/alertas.html', {'alertas_stock': alertas_stock})


# Vista de reportes (solo accesible por Administradores de Obra)
@login_required
@user_passes_test(lambda u: has_role_id(u, ADMINISTRADOR_OBRA), login_url='/access_denied/')
def reports_view(request):
    if request.method == 'POST':
        report_type = request.POST.get('reportType')
        start_date = request.POST.get('startDate')
        end_date = request.POST.get('endDate')

        # Validar y convertir las fechas ingresadas
        try:
            if start_date:
                start_date = make_aware(datetime.strptime(start_date, '%d-%m-%Y'))
            if end_date:
                end_date = make_aware(datetime.strptime(end_date, '%d-%m-%Y'))
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha incorrecto'}, status=400)

        # Manejo para "Inventario Actual"
        if report_type == 'Inventario Actual':
            materiales = Material.objects.select_related('unidad_medida').values(
                'id', 'nombre', 'descripcion', 'cantidad_disponible', 'stock', 'unidad_medida__descripcion', 'activo'
            )
            return JsonResponse({'reportData': list(materiales)})

        # Manejo para "Alertas de Stock Bajo"
        elif report_type == 'Alertas de Stock bajo':
            materiales = Material.objects.select_related('unidad_medida').values(
                'nombre', 'descripcion', 'cantidad_disponible', 'stock', 'unidad_medida__descripcion'
            )
            report_data = [
                {
                    'material': material['nombre'],
                    'descripcion': material['descripcion'],
                    'cantidad_disponible': material['cantidad_disponible'],
                    'stock': material['stock'],
                    'unidad_medida': material['unidad_medida__descripcion'],
                    'alerta_stock': material['cantidad_disponible'] < material['stock']
                }
                for material in materiales if material['cantidad_disponible'] < material['stock']
            ]
            return JsonResponse({'reportData': report_data})

        # Manejo para "Movimientos de Stock"
        elif report_type == 'Movimientos de stock':
            tickets = Ticket.objects.filter(
                estado='cobrado',
                fecha_creacion__range=[start_date, end_date]
            ).values('material_solicitado__nombre', 'cantidad', 'estado', 'fecha_creacion')

            report_data = [
                {
                    'material': ticket['material_solicitado__nombre'],
                    'cantidad': ticket['cantidad'],
                    'estado': ticket['estado'],
                    'fecha_creacion': ticket['fecha_creacion'].strftime('%d-%m-%Y')
                }
                for ticket in tickets
            ]
            return JsonResponse({'reportData': report_data})

    return render(request, 'Modulo_usuario/ReportsView/reports.html')
def user_login(request):
    form = AuthenticationForm(request, data=request.POST or None)
    
    # Verificar si el usuario acaba de cerrar sesión
    logout_message = request.COOKIES.get('logout_success', False)
    show_logout_message = False

    if logout_message:
        show_logout_message = True

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Verificar si el nombre de usuario existe
        if not CustomUser.objects.filter(username=username).exists():
            messages.error(request, 'El usuario ingresado no existe.')
        else:
            # Validar la autenticación del usuario
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                
                # Mostrar el mensaje de éxito solo si no proviene de un cierre de sesión
                if not show_logout_message:
                    messages.success(request, 'Inicio de sesión exitoso.')

                # Redirigir según el rol del usuario
                if has_role_id(user, ADMINISTRADOR_SISTEMA):
                    return redirect('admin_user_list')
                else:
                    return redirect('home')
            else:
                messages.error(request, 'Contraseña incorrecta.')

    response = render(request, 'Modulo_usuario/usuarios/login.html', {
        'form': form,
        'show_logout_message': show_logout_message
    })

    # Limpiar la cookie de logout después de mostrarla
    if show_logout_message:
        response.delete_cookie('logout_success')

    return response


# Crear usuario
# En views.py del administrador de sistema
def create_user(request):
    if request.method == 'POST':
        form = CustomUserForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario creado con éxito.")
            return redirect('Modulo_administrador/usuarios/login_admin.html')  # Cambia esto por el nombre de la vista a la que deseas redirigir
    else:
        form = CustomUserForm()
    
    return render(request, 'Modulo_administrador/usuarios/create_user.html', {'form': form})  # Asegúrate de cambiar 'tu_template.html' por el nombre real de tu plantilla
# Cerrar sesión y redirigir al login
from django.contrib import messages

def logout_view(request):
    logout(request)
    messages.info(request, "Has cerrado sesión exitosamente.")
    return redirect('login')  # Redirige a la página de inicio de sesión

# Vista de acceso denegado
def access_denied_view(request):
    return render(request, 'Modulo_usuario/errors/access_denied.html')


# Crear roles predeterminados
def create_roles(request):
    roles = [
        {'id': ADMINISTRADOR_SISTEMA, 'name': 'Administrador de sistema'},
        {'id': ADMINISTRADOR_OBRA, 'name': 'Administrador de obra'},
        {'id': JEFE_OBRA, 'name': 'Jefe de obra'},
        {'id': CAPATAZ, 'name': 'Capataz'},
        {'id': JEFE_BODEGA, 'name': 'Jefe de Bodega'},
    ]
    
    for role_data in roles:
        Role.objects.get_or_create(id=role_data['id'], defaults={'name': role_data['name']})
        pass


# Vista para listar usuarios con sus roles en el menú de administrador
@login_required
@user_passes_test(lambda u: has_role_id(u, ADMINISTRADOR_SISTEMA), login_url='/access_denied/')
def admin_user_list(request):
    users = CustomUser.objects.prefetch_related('roles').all()  # Asegúrate de que 'roles' sea la relación correcta
    context = {
        'users': users
    }
    return render(request, 'Modulo_administrador/menu_administrador/user_list.html', context)
def lista_usuarios(request):
    users = CustomUser.objects.all()
    return render(request, 'Modulo_administrador/menu_administrador/user_list.html', {'users': users})

@login_required
def inactivar_usuario(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    user.is_active = False
    user.save()
    return redirect('lista_usuarios')

@login_required
def editar_usuario(request, user_id):
    user = CustomUser.objects.get(id=user_id)
    if request.method == 'POST':
        form = CustomUserForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save(commit=False)  # Guarda el usuario sin comprometer el commit todavía
            user.save()  # Guarda el usuario
            form.save_m2m()  # Guarda los roles ManyToMany
            return redirect('lista_usuarios')
        else:
            print(form.errors)  # Imprime errores del formulario si hay alguno
    else:
        form = CustomUserForm(instance=user)
    return render(request, 'Modulo_administrador/menu_administrador/editar_user.html', {'form': form})
@login_required
def crear_usuario(request):
    if request.method == 'POST':
        form = CustomUserForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_usuarios')
    else:
        form = CustomUserForm()
    return render(request, 'Modulo_administrador/usuarios/create_user.html', {'form': form})

@login_required
def activar_usuario(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    user.is_active = True
    user.save()
    return redirect('lista_usuarios')

class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer

class ChoiceViewSet(viewsets.ModelViewSet):
    queryset = Choice.objects.all()
    serializer_class = ChoiceSerializer

@login_required
@user_passes_test(lambda u: has_role_id(u, JEFE_BODEGA), login_url='/access_denied/')
def buscar_material_ajax(request):
    """
    Vista AJAX para buscar materiales por nombre desde la API externa.
    """
    query = request.GET.get('q', '').lower()
    api_url = f"{settings.API_BASE_URL}/materiales/"

    print(f"Consulta de búsqueda: {query}")  # Para depuración

    try:
        response = requests.get(api_url, timeout=5)
        response.raise_for_status()  # Esto lanzará un error si la respuesta no es 200 OK
        materiales_json = response.json()

        print(f"Datos de la API: {materiales_json}")  # Para depuración

        # Filtrar los materiales que coinciden con la consulta
        materiales_filtrados = [
            {
                'id': material['id'],
                'nombre': material['nombre'],
                'unidad_medida': material['unidad_medida'],
                'cantidad_disponible': material['cantidad_disponible']
            }
            for material in materiales_json if query in material['nombre'].lower()
        ]

        print(f"Materiales filtrados: {materiales_filtrados}")  # Para depuración

        return JsonResponse({'materiales': materiales_filtrados})
    
    except requests.exceptions.RequestException as e:
        print(f"Error al obtener datos de la API: {e}")
        return JsonResponse({'error': 'No se pudo obtener la lista de materiales desde la API'}, status=500)
    

def materiales_list(request):
    query = request.GET.get('search', '').strip().lower()

    # Filtrar los materiales que comienzan con el término de búsqueda
    if query:
        materiales = Material.objects.filter(nombre__istartswith=query, activo=True)
    else:
        materiales = Material.objects.none()  # No devolver nada si no hay consulta

    # Convertir los resultados a JSON
    materiales_filtrados = [
        {
            'id': material.id,
            'nombre': material.nombre,
            'unidad_medida': material.unidad_medida.descripcion,
            'cantidad_disponible': material.cantidad_disponible
        }
        for material in materiales
    ]

    return JsonResponse(materiales_filtrados, safe=False)

def format_date(date_str):
    """Convierte una fecha en formato d-m-Y a un objeto datetime."""
    try:
        return datetime.strptime(date_str, "%d-%m-%Y")
    except ValueError:
        return None
@login_required
def export_to_pdf(request):
    report_type = request.GET.get('reportType', None)
    start_date_str = request.GET.get('startDate', None)
    end_date_str = request.GET.get('endDate', None)

    # Validar los parámetros
    if not report_type or not start_date_str or not end_date_str:
        return JsonResponse({'error': 'Faltan parámetros para generar el reporte'}, status=400)

    # Convertir fechas al formato correcto
    start_date = format_date(start_date_str)
    end_date = format_date(end_date_str)
    if not start_date or not end_date:
        return JsonResponse({'error': 'Formato de fecha incorrecto'}, status=400)

    # Crear el archivo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={report_type}_report.pdf'
    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Encabezado del PDF
    pdf.setFont("Helvetica-Bold", 16)
    pdf.setFillColor(colors.HexColor("#007bff"))
    pdf.drawString(2 * cm, height - 2 * cm, f"📊 Reporte de {report_type}")
    pdf.setFont("Helvetica", 12)
    pdf.setFillColor(colors.black)
    pdf.drawString(2 * cm, height - 3 * cm, f"Desde: {start_date_str}  Hasta: {end_date_str}")

    # Línea de separación
    pdf.setStrokeColor(colors.HexColor("#007bff"))
    pdf.setLineWidth(1)
    pdf.line(2 * cm, height - 3.5 * cm, width - 2 * cm, height - 3.5 * cm)

    # Contenido según el tipo de reporte
    y_position = height - 5 * cm
    pdf.setFont("Helvetica", 11)

    if report_type == 'Inventario Actual':
        pdf.setFillColor(colors.HexColor("#4caf50"))
        pdf.drawString(2 * cm, y_position, "ID    |    Nombre    |    Descripción    |    Unidad de Medida    |    Cantidad    |    Estado")
        y_position -= 1 * cm
        materiales = Material.objects.all()
        for material in materiales:
            estado = 'Activo' if material.activo else 'Inactivo'
            pdf.setFillColor(colors.black)
            pdf.drawString(2 * cm, y_position, f"{material.id}  |  {material.nombre}  |  {material.descripcion}  |  {material.unidad_medida.descripcion}  |  {material.cantidad_disponible}  |  {estado}")
            y_position -= 0.8 * cm
            if y_position < 2 * cm:
                pdf.showPage()
                y_position = height - 3 * cm

    elif report_type == 'Alertas de Stock bajo':
        pdf.setFillColor(colors.HexColor("#e53935"))
        pdf.drawString(2 * cm, y_position, "Material    |    Cantidad Disponible    |    Stock Mínimo    |    Alerta")
        y_position -= 1 * cm
        materiales = Material.objects.filter(cantidad_disponible__lt=models.F('stock'))
        for material in materiales:
            alerta = 'Stock Bajo' if material.cantidad_disponible < material.stock else 'Ok'
            pdf.setFillColor(colors.black)
            pdf.drawString(2 * cm, y_position, f"{material.nombre}  |  {material.cantidad_disponible}  |  {material.stock}  |  {alerta}")
            y_position -= 0.8 * cm
            if y_position < 2 * cm:
                pdf.showPage()
                y_position = height - 3 * cm

    elif report_type == 'Movimientos de stock':
        pdf.setFillColor(colors.HexColor("#3f51b5"))
        pdf.drawString(2 * cm, y_position, "Material    |    Cantidad    |    Estado    |    Fecha")
        y_position -= 1 * cm
        tickets = Ticket.objects.filter(
            estado='cobrado',
            fecha_creacion__range=[start_date, end_date]
        )
        for ticket in tickets:
            pdf.setFillColor(colors.black)
            pdf.drawString(2 * cm, y_position, f"{ticket.material_solicitado.nombre}  |  {ticket.cantidad}  |  {ticket.estado}  |  {ticket.fecha_creacion.strftime('%d-%m-%Y')}")
            y_position -= 0.8 * cm
            if y_position < 2 * cm:
                pdf.showPage()
                y_position = height - 3 * cm

    # Pie de página
    pdf.setFillColor(colors.HexColor("#007bff"))
    pdf.setFont("Helvetica-Oblique", 10)
    pdf.drawString(2 * cm, 1.5 * cm, "Generado por el Sistema de Gestión de Inventario - Empresa Constructora")

    # Guardar y cerrar el PDF
    pdf.save()
    return response

@login_required
def export_to_excel(request):
    # Obtener los parámetros del formulario desde la URL
    report_type = request.GET.get('reportType')
    start_date = request.GET.get('startDate')
    end_date = request.GET.get('endDate')

    # Crear un archivo de Excel y una hoja activa
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = report_type

    # Configurar el nombre del archivo descargable
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={report_type}_report.xlsx'

    # Convertir las fechas utilizando la función auxiliar
    start_date = format_date(start_date)
    end_date = format_date(end_date)

    # Validar que las fechas sean válidas
    if not start_date or not end_date:
        sheet.append(['Formato de fecha incorrecto'])
        workbook.save(response)
        return response

    # Estilos para el encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Definir encabezados según el tipo de reporte
    if report_type == 'Inventario Actual':
        headers = ['ID', 'Nombre', 'Descripción', 'Unidad de Medida', 'Cantidad Disponible', 'Stock Mínimo', 'Estado']
        materiales = Material.objects.all()

        # Agregar encabezados
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center

        # Agregar datos
        for material in materiales:
            estado = 'Activo' if material.activo else 'Inactivo'
            row = [
                material.id,
                material.nombre,
                material.descripcion,
                material.unidad_medida.descripcion,
                material.cantidad_disponible,
                material.stock,
                estado
            ]
            sheet.append(row)

    elif report_type == 'Alertas de Stock bajo':
        headers = ['Material', 'Cantidad Disponible', 'Stock Mínimo', 'Unidad de Medida', 'Alerta']
        materiales = Material.objects.filter(cantidad_disponible__lt=models.F('stock'))

        # Agregar encabezados
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center

        # Agregar datos
        for material in materiales:
            alerta = 'Stock Bajo' if material.cantidad_disponible < material.stock else 'Ok'
            row = [
                material.nombre,
                material.cantidad_disponible,
                material.stock,
                material.unidad_medida.descripcion,
                alerta
            ]
            sheet.append(row)

    elif report_type == 'Movimientos de stock':
        headers = ['Material', 'Cantidad', 'Estado', 'Fecha de Creación']
        tickets = Ticket.objects.filter(
            estado='cobrado',
            fecha_creacion__range=[start_date, end_date]
        )

        # Agregar encabezados
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center

        # Agregar datos
        for ticket in tickets:
            row = [
                ticket.material_solicitado.nombre,
                ticket.cantidad,
                ticket.estado,
                ticket.fecha_creacion.strftime('%d-%m-%Y')
            ]
            sheet.append(row)

    else:
        sheet.append(['Tipo de reporte no válido'])

    # Ajustar el ancho de las columnas
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = max_length

    # Guardar el archivo Excel
    workbook.save(response)
    return response

@login_required
@user_passes_test(lambda u: u.is_authenticated)
def movimientos_view(request):
    # Obtener todos los movimientos de stock (tickets) de la base de datos
    movimientos = Ticket.objects.select_related('material_solicitado').order_by('-fecha_creacion')

    # Filtros por fecha si se envían en la solicitud
    start_date = request.GET.get('startDate')
    end_date = request.GET.get('endDate')

    if start_date:
        start_date = make_aware(datetime.strptime(start_date, '%d-%m-%Y'))
        movimientos = movimientos.filter(fecha_creacion__gte=start_date)
    if end_date:
        end_date = make_aware(datetime.strptime(end_date, '%d-%m-%Y'))
        movimientos = movimientos.filter(fecha_creacion__lte=end_date)

    context = {
        'movimientos': movimientos,
        'start_date': start_date,
        'end_date': end_date
    }
    return render(request, 'Modulo_usuario/ReportsView/movimientos.html', context)